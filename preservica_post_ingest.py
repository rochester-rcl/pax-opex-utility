import os
import os.path
import csv
import time
from datetime import datetime
from pyrsistent import thaw
from os.path import basename
from pyPreservica import *
from deepdiff import DeepDiff
from openpyxl import load_workbook

mincol = 2
maxcol = 2
minrow = 2
refidcol = 2
aouricol = 3
titlecol = 7
datecol = 12
filecol = 13
prescol = 14

def test_connection(window, mline, alt_background, init_color, update_color, summary_color, user_name, pass_word, ten_ancy, ser_ver, two_factorcb, two_factorkey):
    mline.update('[START] TESTING CONNECTION TO PRESERVICA\n', append=True, text_color_for_value=init_color)
    window.refresh()
    if two_factorcb == True:
        client = EntityAPI(username=user_name, password=pass_word, tenant=ten_ancy, server=ser_ver, two_fa_secret_key=two_factorkey)
    else:
        client = EntityAPI(username=user_name, password=pass_word, tenant=ten_ancy, server=ser_ver)
    root_folders = client.children(None)
    root_folders = str(client.children(None))
    mline.update('[SUMMARY]\n' + root_folders + '\n\n', append=True, text_color_for_value=update_color)
    window.refresh()
    
def move_opex_aspace(window, mline, alt_background, init_color, update_color, summary_color, user_name, pass_word, ten_ancy, ser_ver, two_factorcb, two_factorkey, opex_folder, aspace_folder):
    mline.update('[START] MOVING ASSETS TO PENDING LINK\n', append=True, text_color_for_value=init_color)
    window.refresh()
    if two_factorcb == True:
        client = EntityAPI(username=user_name, password=pass_word, tenant=ten_ancy, server=ser_ver, two_fa_secret_key=two_factorkey)
    else:
        client = EntityAPI(username=user_name, password=pass_word, tenant=ten_ancy, server=ser_ver)
    opex_folder = client.descendants(opex_folder)
    aspace_folder = client.folder(aspace_folder)
    count = 0
    for entity in opex_folder:
        client.move(entity, aspace_folder)
        count += 1
        if (count % 2) == 0:
            mline.update('[UPDATE] moving item {}\n'.format(str(count)), append=True, text_color_for_value=update_color)
            window.refresh()
        else:
            mline.update('[UPDATE] moving item {}\n'.format(str(count)), append=True, text_color_for_value=update_color, background_color_for_value=alt_background)
            window.refresh()
        time.sleep(1)
    mline.update('[SUMMARY] moved {} entities\n\n'.format(str(count)), append=True, text_color_for_value=summary_color)
    window.refresh()

def move_aspace_trash(window, mline, alt_background, init_color, update_color, summary_color, user_name, pass_word, ten_ancy, ser_ver, two_factorcb, two_factorkey, aspace_folder, trash_folder):
    mline.update('[START] MOVING EMPTY FOLDERS TO TRASH\n', append=True, text_color_for_value=init_color)
    window.refresh()
    if two_factorcb == True:
        client = EntityAPI(username=user_name, password=pass_word, tenant=ten_ancy, server=ser_ver, two_fa_secret_key=two_factorkey)
    else:
        client = EntityAPI(username=user_name, password=pass_word, tenant=ten_ancy, server=ser_ver)
    aspace_folder = client.descendants(aspace_folder)
    count = 0
    now = datetime.now()
    folder_title = now.strftime('%Y-%m-%d_%H-%M-%S') + '_Deletion'
    new_folder = client.create_folder(folder_title, "container folder to delete AO# folders", 'closed', trash_folder)
    dest_folder = client.folder(new_folder.reference)
    for entity in aspace_folder:
        test_var = entity.title
        if test_var.startswith('archival_object_'):
            client.move(entity, dest_folder)
            count += 1
            if (count % 2) == 0:
                mline.update('[UPDATE] moving item {}\n'.format(str(count)), append=True, text_color_for_value=update_color)
                window.refresh()
            else:
                mline.update('[UPDATE] moving item {}\n'.format(str(count)), append=True, text_color_for_value=update_color, background_color_for_value=alt_background)
                window.refresh()
            time.sleep(1)
    mline.update('[SUMMARY] Moved {} folders into the trash\n\n'.format(str(count)), append=True, text_color_for_value=summary_color)
    window.refresh()

def quality_control(window, mline, alt_background, init_color, update_color, summary_color, user_name, password, tenancy, server, twofactorcb, twofactorkey, qual_control, workorder, worksheet, maxrow):
    mline.update('[START] STARTING QA\n', append=True, text_color_for_value=init_color)
    window.refresh()
    asset_count = 0
    mline.update('[UPDATE] MAKING DROID DICTIONARY\n', append=True, text_color_for_value=update_color, background_color_for_value=alt_background)
    window.refresh()
    manifestdict = dict()
    with open(qual_control, newline = '') as csvfile:
        reader = csv.reader(csvfile, delimiter = ',', quotechar = '"')
        next(reader)
        for row in reader: 
                manifestdict[row[0]] = row[1]
    mline.update('[UPDATE] MAKING PRESERVICA DICTIONARY\n', append=True, text_color_for_value=update_color)
    window.refresh()
    if twofactorcb == True:
        client = EntityAPI(username=user_name, password=password, tenant=tenancy, server=server, two_fa_secret_key=twofactorkey)
    else:
        client = EntityAPI(username=user_name, password=password, tenant=tenancy, server=server)
    wb = load_workbook(workorder)
    ws = wb[worksheet]
    maximumrow = int(maxrow)
    iterrow = 2
    ws.cell(row = 1, column = prescol).value = 'Preservica UUID'
    for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maximumrow, max_col = maxcol):
        for cell in row:
            ref_id = ws.cell(row = iterrow, column = refidcol).value
            for ident in filter(only_assets, client.identifier("code", ref_id)):
                asset = client.asset(ident.reference)
                ws.cell(row = iterrow, column = prescol).value = ident.reference
    wb.save(workorder)
    preservicalist = list()
    iterrow = 2
    for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maximumrow, max_col = maxcol):
        for cell in row:
            preservicalist.append(ws.cell(row = iterrow, column = prescol).value)
        iterrow += 1
    preservicadict = dict()
    for reference in preservicalist:
        asset = client.asset(reference)
        asset_count += 1
        for representation in client.representations(asset):
            for content_object in client.content_objects(representation):
                for generation in client.generations(content_object):
                    for bitstream in generation.bitstreams:
                        filename = bitstream.filename 
                        for algorithm,value in bitstream.fixity.items():
                            preservicadict[filename] = value    
    mline.update('[UPDATE] COMPARING DICTIONARIES\n', append=True, text_color_for_value=update_color, background_color_for_value=alt_background)
    window.refresh()
    diff = DeepDiff(preservicadict, manifestdict, verbose_level=2)
    if len(diff) == 0:
        mline.update('[SUMMARY] QUALITY ASSURANCE PASSED\n\n', append=True, text_color_for_value=summary_color)
        window.refresh()
    else:
        mline.update('[SUMMARY] ERROR ENCOUNTERED\n', append=True, text_color_for_value=summary_color)
        mline.update('REVIEW OUTPUT BELOW FOR PROBLEMS\n', append=True, text_color_for_value=summary_color)
        diff_str = str(diff)
        mline.update(diff_str + '\n', append=True, text_color_for_value=summary_color)
        window.refresh()
