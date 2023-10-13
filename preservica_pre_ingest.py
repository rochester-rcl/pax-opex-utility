import os
import shutil
import pathlib
import hashlib
import csv
import time
from datetime import datetime
from pyrsistent import thaw
from zipfile import ZipFile
from os.path import basename
from cleanup_dates import *
from openpyxl import load_workbook

mincol = 2
maxcol = 2
minrow = 2
refidcol = 2
aouricol = 3
titlecol = 7
datecol = 12
filecol = 13

container = ''

#this function takes the folder containing all the preservation masters and renames to be the "container" folder which will ultimately be used for OPEX incremental ingest
#also creates a "project_log.txt" file to store variables so that an ingest project can be worked on over multiple sessions
def create_container(window, mline, projpath, work_order):
    mline.update('----CREATING CONTAINER----\n', text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    now = datetime.now()
    date_time = now.strftime('%Y-%m-%d_%H-%M-%S')
    global container
    container = 'container_' + date_time
    os.mkdir(os.path.join(projpath, container))
    workorder = os.path.basename(work_order)
    for file in os.listdir(path = projpath):
        if file == workorder or file == container:
            continue
        else:
            shutil.move(os.path.join(projpath, file), os.path.join(projpath, container, file))
    mline.update('Container directory: {} and moved digital assets into it\n'.format(container), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

def file_hash_list(window, mline, projpath):
    mline.update('----CREATING LIST OF FILES AND CHECKSUMS----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    path_container = os.path.join(projpath, container)
    file_count = 0
    with open(os.path.join(projpath, container + '.csv'), mode='w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        for file in os.listdir(path = path_container):
            file_hand = open(os.path.join(path_container, file), 'rb')
            file_read = file_hand.read()
            sha1_checksum = hashlib.sha1(file_read).hexdigest()
            csv_writer.writerow([file, sha1_checksum])
            file_count += 1
            if (file_count % 2) == 0:
                mline.update('file: {} checksum: {}\n'.format(file, sha1_checksum), append=True, text_color_for_value='black', background_color_for_value='white')
                window.refresh()
            else:
                mline.update('file: {} checksum: {}\n'.format(file, sha1_checksum), append=True, text_color_for_value='white', background_color_for_value='black')
                window.refresh()
    mline.update('Generated checksums for {} files\n'.format(file_count), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

#This function creates paths to the folders and files and then moves the files to their respective folders.
def folder_ds_files(window, mline, projpath, filename_delimiter):
    mline.update('----CREATING FOLDER STRUCTURE FOR PRESERVATION MASTERS----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    folder_count = 0
    file_count = 0
    loop_count = 0
    path_container = os.path.join(projpath, container)
    folder_list = list()
    for file in os.listdir(path = path_container):
        file_root = file.split('.')[0].strip()
        if filename_delimiter in file_root:
            file_root = file_root.split(filename_delimiter)[0].strip()
        if file_root not in folder_list:
            folder_list.append(file_root)
            loop_count += 1
            if (loop_count % 2) == 0:
                mline.update('added {} to folder_list\n'.format(file_root), append=True, text_color_for_value='black', background_color_for_value='white')
                window.refresh()
            else:
                mline.update('added {} to folder_list\n'.format(file_root), append=True, text_color_for_value='white', background_color_for_value='black')
                window.refresh()
    for file_root in folder_list:
        path_folder = os.path.join(path_container, file_root)
        os.mkdir(path_folder)
        loop_count += 1
        if (loop_count % 2) == 0:
            mline.update('created {}\n'.format(path_folder), append=True, text_color_for_value='black', background_color_for_value='white')
            window.refresh()
        else:
            mline.update('created {}\n'.format(path_folder), append=True, text_color_for_value='white', background_color_for_value='black')
            window.refresh()
        folder_count += 1
    for file in os.listdir(path = path_container):
        if '.' not in file:
            continue
        else:
            path_file = os.path.join(path_container, file)
            file_prefix = file.split('.')[0].strip()
            if '-' in file_prefix:
                file_prefix = file_prefix.split('-')[0].strip()
            path_folder = os.path.join(path_container, file_prefix, file)
            shutil.move(path_file, path_folder)
            loop_count += 1
            if (loop_count % 2) == 0:
                mline.update('moved {} to {}\n'.format(path_file, path_folder), append=True, text_color_for_value='black', background_color_for_value='white')
                window.refresh()
            else:
                mline.update('moved {} to {}\n'.format(path_file, path_folder), append=True, text_color_for_value='white', background_color_for_value='black')
                window.refresh()
            file_count += 1
    mline.update('created {} folders\n'.format(folder_count), append=True, text_color_for_value='white', background_color_for_value='green')
    mline.update('moved {} files\n'.format(file_count), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

#this function begins to create the PAX structure
#putting PDFs in Representation_Access folders and TIFFs in Representation_Preservation folders
def representation_preservation_access(window, mline, projpath, rep_pres, rep_acc):
    mline.update('----CREATING REPRESENTATION FOLDERS AND MOVING ASSETS INTO THEM----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    folder_count = 0
    file_count = 0
    loop_count = 0
    path_container = os.path.join(projpath, container)
    rep_pres_name = 'Representation_Preservation'
    rep_acc_name = 'Representation_Access'
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(projpath, container, directory)
        path_pres = os.path.join(projpath, container, directory, rep_pres_name)
        path_acc = os.path.join(projpath, container, directory, rep_acc_name)
        os.mkdir(path_pres)
        os.mkdir(path_acc)
        folder_count += 2
        for file in os.listdir(path = path_directory):
            path_directoryfile = os.path.join(projpath, container, directory, file)
            if file == rep_pres_name or file == rep_acc_name:
                continue
            if file.endswith(rep_acc):
                file_name = file.split('.')[0]
                os.mkdir(os.path.join(path_acc, file_name))
                shutil.move(path_directoryfile, os.path.join(path_acc, file_name, file))
                loop_count += 1
                if (loop_count % 2) == 0:
                    mline.update('created directory: {}\n'.format(path_acc + '/' + file_name), append=True, text_color_for_value='black', background_color_for_value='white')
                    mline.update('moved file: {}\n'.format(path_acc + '/' + file_name + '/' + file), append=True, text_color_for_value='black', background_color_for_value='white')
                    window.refresh()
                else:
                    mline.update('created directory: {}\n'.format(path_acc + '/' + file_name), append=True, text_color_for_value='white', background_color_for_value='black')
                    mline.update('moved file: {}\n'.format(path_acc + '/' + file_name + '/' + file), append=True, text_color_for_value='white', background_color_for_value='black')
                    window.refresh()
                file_count += 1
                window.refresh()
            if file.endswith(rep_pres):
                file_name = file.split('.')[0]
                os.mkdir(os.path.join(path_pres, file_name))
                shutil.move(path_directoryfile, os.path.join(path_pres, file_name, file))
                loop_count += 1
                if (loop_count % 2) == 0:
                    mline.update('created directory: {}\n'.format(path_pres + '/' + file_name), append=True, text_color_for_value='black', background_color_for_value='white')
                    mline.update('moved file: {}\n'.format(path_pres + '/' + file_name + '/' + file), append=True, text_color_for_value='black', background_color_for_value='white')
                    window.refresh()
                else:
                    mline.update('created directory: {}\n'.format(path_pres + '/' + file_name), append=True, text_color_for_value='white', background_color_for_value='black')
                    mline.update('moved file: {}\n'.format(path_pres + '/' + file_name + '/' + file), append=True, text_color_for_value='white', background_color_for_value='black')
                    window.refresh()
                file_count += 1
                window.refresh()
    mline.update('Created {} Representation directories | Moved {} files into created directories\n'.format(folder_count, file_count), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

#this function stages the "Representation_" folders for each asset inside a new directory,
#then zipes the files into individual PAX objects and deletes the source files
def create_pax(window, mline, projpath):
    mline.update('----CREATING PAX OBJECTS----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    pax_count = 0
    path_container = os.path.join(projpath, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(projpath, container, directory)
        path_paxstage = os.path.join(projpath, container, directory, 'pax_stage')
        os.mkdir(path_paxstage)
        shutil.move(os.path.join(path_directory, 'Representation_Preservation'), path_paxstage)
        shutil.move(os.path.join(path_directory, 'Representation_Access'), path_paxstage)
        path_directory = os.path.join(projpath, container, directory)
        zip_dir = pathlib.Path(path_paxstage)
        pax_obj = ZipFile(os.path.join(path_directory, directory + '.zip'), 'w')
        for file_path in zip_dir.rglob("*"):
            pax_obj.write(file_path, arcname = file_path.relative_to(zip_dir))
        pax_obj.close()
        os.replace(os.path.join(path_directory, directory + '.zip'), os.path.join(path_directory, directory + '.pax.zip'))
        time.sleep(1)
        pax_count += 1
        shutil.rmtree(path_paxstage)
        if (pax_count % 2) == 0:
            mline.update('created {}\n'.format(str(pax_count) + ': ' + directory + '.pax.zip'), append=True, text_color_for_value='black', background_color_for_value='white')
            window.refresh()
        else:
            mline.update('created {}\n'.format(str(pax_count) + ': ' + directory + '.pax.zip'), append=True, text_color_for_value='white', background_color_for_value='black')
            window.refresh()
    mline.update('Created {} PAX objects\n'.format(pax_count), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

#this function creates the OPEX metadata file that accompanies an individual zipped PAX package
#this function also includes the metadata necessary for ArchivesSpace sync to Preservica
def pax_metadata(window, mline, projpath, workorder, worksheet, maxrow, format_dates):
    mline.update('---CREATING METADATA FILES FOR PAX OBJECTS----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    wb = load_workbook(workorder)
    ws = wb[worksheet]
    dir_count = 0
    maximumrow = int(maxrow)
    path_container = os.path.join(projpath, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(projpath, container, directory)
        pax_hand = open(os.path.join(path_directory, directory + '.pax.zip'), 'rb')
        pax_read = pax_hand.read()
        sha1_checksum = hashlib.sha1(pax_read).hexdigest()
        pax_hand.close()
        iterrow = 2
        for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maximumrow, max_col = maxcol):
            for cell in row:
                cuid = ws.cell(row = iterrow, column = filecol).value
                if cuid == directory:
                    ref_id = ws.cell(row = iterrow, column = refidcol).value
                    title = ws.cell(row = iterrow, column = titlecol).value
                    if '&' in title:
                        title = title.replace('&', 'and')
                    if format_dates == True:
                        date_full = ws.cell(row = iterrow, column = datecol).value
                        date_formatted = aspace_dates(date_full)
                        display_title = '{title}{date_formatted}'.format(title=title, date_formatted=date_formatted)
                    else:
                        display_title = '{title}'.format(title=title)
                    opex = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0">
    <opex:Transfer>
        <opex:Fixities>
            <opex:Fixity type="SHA-1" value="{sha1_checksum}"/>
        </opex:Fixities>
    </opex:Transfer>
    <opex:Properties>
        <opex:Title>{title}</opex:Title>
        <opex:Identifiers>
            <opex:Identifier type="code">{ref_id}</opex:Identifier>
        </opex:Identifiers>
    </opex:Properties>
    <opex:DescriptiveMetadata>
        <LegacyXIP xmlns="http://preservica.com/LegacyXIP">
            <AccessionRef>catalogue</AccessionRef>
        </LegacyXIP>
    </opex:DescriptiveMetadata>
</opex:OPEXMetadata>'''.format(sha1_checksum=sha1_checksum, title=display_title, ref_id=ref_id)
                    filename = directory + '.pax.zip.opex'
                    pax_md_hand = open(os.path.join(path_directory, filename), 'w')
                    pax_md_hand.write(opex)
                    pax_md_hand.close()
                    dir_count += 1
                    if (dir_count % 2) == 0:
                        mline.update('created {}\n'.format(filename), append=True, text_color_for_value='black', background_color_for_value='white')
                        window.refresh()
                    else:
                        mline.update('created {}\n'.format(filename), append=True, text_color_for_value='white', background_color_for_value='black')
                        window.refresh()
            iterrow += 1
    mline.update('Created {} OPEX metdata files for individual assets\n'.format(dir_count), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

#this function matches directory names (based on CUID) with archival object numbers from work order spreadsheet
#this metadata is another facet required for ArchivesSpace to Preservica synchronization
def ao_opex_metadata(window, mline, projpath, workorder, worksheet, maxrow):
    mline.update('----CREATE ARCHIVAL OBJECT OPEX METADATA----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    wb = load_workbook(workorder)
    ws = wb[worksheet]
    file_count = 0
    path_container = os.path.join(projpath, container)
    maximumrow = int(maxrow)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(projpath, container, directory)
        iterrow = 2
        for row in ws.iter_rows(min_row = minrow, min_col = mincol, max_row = maximumrow, max_col = maxcol):
            for cell in row:
                cuid = ws.cell(row = iterrow, column = filecol).value
                if cuid == directory:
                    ao_num_uri = ws.cell(row = iterrow, column = aouricol).value
                    ao_num = 'archival_object_' + ao_num_uri.split('/')[-1].strip()
                    opex = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0">
    <opex:Properties>
        <opex:Title>{ao_num}</opex:Title>
        <opex:Identifiers>
            <opex:Identifier type="code">{ao_num}</opex:Identifier>
        </opex:Identifiers>
    </opex:Properties>
    <opex:DescriptiveMetadata>
        <LegacyXIP xmlns="http://preservica.com/LegacyXIP">
            <Virtual>false</Virtual>
        </LegacyXIP>
    </opex:DescriptiveMetadata>
</opex:OPEXMetadata>'''.format(ao_num = ao_num)
                    with open(os.path.join(path_directory, ao_num + '.opex'), 'w') as ao_md:
                        ao_md.write(opex)
                    file_count += 1
                    os.replace(path_directory, os.path.join(path_container, ao_num))
                    time.sleep(1)
                    if (file_count % 2) == 0:
                        mline.update('processed folder metadata in new folder: {}\n'.format(ao_num), append=True, text_color_for_value='black', background_color_for_value='white')
                        window.refresh()
                    else:
                        mline.update('processed folder metadata in new folder: {}\n'.format(ao_num), append=True, text_color_for_value='white', background_color_for_value='black')
                        window.refresh()

            iterrow += 1
    mline.update('Created {} archival object metadata files\n'.format(file_count), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()

#this function creates the last OPEX metadata file required for the OPEX incremental ingest, for the container folder
#this OPEX file has the folder manifest to ensure that content is ingested properly
def opex_container_metadata(window, mline, projpath):
    mline.update('----CREATE CONTAINER OBJECT OPEX METADATA----\n', append=True, text_color_for_value='white', background_color_for_value='red')
    window.refresh()
    opex1 = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0">
    <opex:Transfer>
        <opex:Manifest>
            <opex:Folders>\n'''
    opex2 = ''
    path_container = os.path.join(projpath, container)
    for directory in os.listdir(path = path_container):
        opex2 += '\t\t\t\t<opex:Folder>' + directory + '</opex:Folder>\n'
    opex3 = '''\t\t\t</opex:Folders>
        </opex:Manifest>
    </opex:Transfer>
</opex:OPEXMetadata>'''
    container_opex_hand = open(os.path.join(projpath, container, container + '.opex'), 'w')
    container_opex_hand.write(opex1 + opex2 + opex3)
    mline.update('Created OPEX metadata file for {} directory\n'.format(container), append=True, text_color_for_value='white', background_color_for_value='green')
    window.refresh()
    container_opex_hand.close()
